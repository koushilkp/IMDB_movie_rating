# first create venv >> and activate
# pip install requests
# pip install bs4
# pip install html5lib

import pandas as pd
import requests
import openpyxl
from bs4 import BeautifulSoup

# to load in the xsxl file  we need to import openpyxl file and check the  the no. of sheet with name 

excel=openpyxl.Workbook()   
# print(excel.sheetnames)
sheet=excel.active

sheet.title ='Top rated Moives'
print(excel.sheetnames)

# here we will create column name for excel
sheet.append(["Movie Rank","Movie name","year of Release","IMDB rating"])


try:
    source=requests.get('https://www.imdb.com/chart/top/')
    source.raise_for_status()  # of the ther error 404 the if the web is present or not

    soup=BeautifulSoup(source.text,'html.parser')
    # print(soup)
    movies = soup.find('tbody',class_='lister-list').find_all('tr')
    # print(len(movies))   #250
    # print(movies)          #this will give list

    for movie in movies:

        name=movie.find('td', class_='titleColumn').a.text
        # print("Movie_name>>>>>",name)

        rank=movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]
        # print("movie_Rank>>>>>>",rank)

        year=movie.find('td',class_='titleColumn').find('span',class_="secondaryInfo").text.strip('()')
        # print("Movie_year>>>>>>",year)

        rating=movie.find('td',class_='ratingColumn imdbRating').strong.text
        # print("Movie_ratting>>>>>>",rating)

        print(rank,name,year,rating)
        sheet.append([rank,name,year,rating])

    
except Exception as e :
    print(e)

excel.save("IMDb top 250 movie rating.xlsx")